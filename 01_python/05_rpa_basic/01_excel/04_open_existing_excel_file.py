"""
📌 Reading Data from an Existing Excel File (openpyxl - load_workbook)

This script demonstrates how to:
- Load an existing Excel (.xlsx) file using Python
- Access the active worksheet
- Read cell values without opening Excel itself
- Print Excel data in a row/column structure inside the terminal

Purpose of this file:
- Understand how load_workbook() works
- Learn how to access sheets and cells
- Use nested loops to read Excel data
- Control print output using the `end` parameter
"""

# ==================================================
# 1️⃣ Import required library
# ==================================================
from openpyxl import load_workbook  # Used to load an existing Excel file


# ==================================================
# 2️⃣ Open the Excel file
# ==================================================
# Load the existing Excel file (must be in the same directory)
wb = load_workbook("sample.xlsx")

# Get the currently active worksheet
ws = wb.active


# ==================================================
# 3️⃣ Read cells when the range is known
# ==================================================
# Assumption:
# - Rows: 1 ~ 10
# - Columns: 1 ~ 10

for row in range(1, 11):
    for col in range(1, 11):
        # ws.cell(row, column).value
        # → Retrieves the actual value stored in the cell
        print(ws.cell(row=row, column=col).value, end=" ")
    print()  # Move to the next line after each row


# ==================================================
# 4️⃣ Read all cells when the range is unknown
# ==================================================
# max_row and max_column automatically detect
# the last row and column containing data

for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        print(ws.cell(row=row, column=col).value, end=" ")
    print()


# ==================================================
# 5️⃣ Why use end=" " in print()?
# ==================================================
"""
By default, print() adds a newline after each call.

Example:
print(1)
print(2)
print(3)

Output:
1
2
3

This results in vertical output, which does NOT match Excel's layout.

Using end=" " replaces the newline with a space,
allowing values to be printed horizontally.
"""

"""
Example:
print(1, end=" ")
print(2, end=" ")
print(3, end=" ")

Output:
1 2 3

This mimics how data appears across a single row in Excel.
"""

"""
In this script:
- print(..., end=" ") → moves horizontally across columns
- print()             → moves down to the next row

Together, they recreate Excel's row/column structure in the terminal.
"""


# ==================================================
# 📌 Summary
# ==================================================
"""
✔ load_workbook("file.xlsx")
   → Loads an existing Excel file into Python

✔ wb.active
   → Selects the active worksheet

✔ ws.cell(row, column).value
   → Reads the value from a specific cell

✔ ws.max_row / ws.max_column
   → Automatically detect the data range

✔ end=" "
   → Prevents automatic line breaks and enables horizontal output

This approach allows Excel data to be inspected
directly from the Python terminal without opening Excel.
"""
