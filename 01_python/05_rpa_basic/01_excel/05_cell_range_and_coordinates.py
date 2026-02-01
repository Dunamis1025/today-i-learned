"""
📌 Working with Rows, Columns, and Cell Coordinates (openpyxl)

This script demonstrates how to:
- Append rows of data to an Excel worksheet
- Generate random score data
- Select specific columns and column ranges
- Select specific rows and row ranges
- Exclude header rows when processing data
- Extract row/column information from cell coordinates

This file is intended as a learning note to build
a solid understanding of how openpyxl handles
row-based and column-based data access.
"""

# ==================================================
# 1️⃣ Import required libraries
# ==================================================
from openpyxl import Workbook
from random import randint
from openpyxl.utils.cell import coordinate_from_string


# ==================================================
# 2️⃣ Create Workbook and Worksheet
# ==================================================
wb = Workbook()
ws = wb.active


# ==================================================
# 3️⃣ Append data row by row
# ==================================================
# Add header row
ws.append(["ID", "English", "Math"])

# Generate sample student score data
# English and Math scores are random values between 0 and 100
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])


# ==================================================
# 4️⃣ Select specific columns
# ==================================================
# ws["B"] → entire column B (English scores)
col_english = ws["B"]

# Example usage:
# for cell in col_english:
#     print(cell.value)


# ==================================================
# 5️⃣ Select multiple columns at once
# ==================================================
# ws["B:C"] → English and Math columns
col_range = ws["B:C"]

# Example usage:
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)


# ==================================================
# 6️⃣ Select specific rows
# ==================================================
# ws[1] → first row (header)
header_row = ws[1]

# Example usage:
# for cell in header_row:
#     print(cell.value)


# ==================================================
# 7️⃣ Select a specific row range (excluding header)
# ==================================================
# Example: rows 2 to 6
#
# row_range = ws[2:6]
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()


# ==================================================
# 8️⃣ Select rows dynamically using max_row
# ==================================================
# ws.max_row automatically detects the last row with data
row_range = ws[2:ws.max_row]  # exclude header row

for rows in row_range:
    for cell in rows:
        # cell.coordinate returns a string like "A1", "B3"
        coord = coordinate_from_string(cell.coordinate)

        # coord[0] → column letter
        # coord[1] → row number
        print(coord[0], end="")
        print(coord[1], end=" ")
    print()


# ==================================================
# 9️⃣ Save Excel file
# ==================================================
wb.save("sample.xlsx")


# ==================================================
# 📌 Summary
# ==================================================
"""
✔ ws.append([...])
   → Add data row by row to a worksheet

✔ ws["B"], ws["B:C"]
   → Select single or multiple columns

✔ ws[1], ws[2:6]
   → Select specific rows or row ranges

✔ ws.max_row
   → Automatically detect the last row containing data

✔ cell.coordinate
   → Get the Excel-style cell reference (e.g. "A1")

✔ coordinate_from_string()
   → Split a cell reference into column letter and row number

This script reinforces core concepts for navigating
Excel data using openpyxl.
"""
