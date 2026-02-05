"""
📌 Unmerging Cells in Excel (openpyxl - unmerge_cells)

This script demonstrates how to UNMERGE cells that were previously merged
in an existing Excel (.xlsx) file using openpyxl.

What this script covers:
- Loading an existing Excel file with merged cells
- Unmerging a specific merged cell range
- Saving the result as a new Excel file (recommended to keep the original)

Key points:
- unmerge_cells() must receive the EXACT range that was merged
- After unmerging, the value remains in the top-left cell
- Other cells in the original merged range will be empty
"""

# ==================================================
# 1️⃣ Import required library
# ==================================================
from openpyxl import load_workbook

# ==================================================
# 2️⃣ Load an existing Excel file
# ==================================================
# The file "sample_merge.xlsx" must exist in the same directory
# This file is assumed to already contain merged cells
wb = load_workbook("sample_merge.xlsx")

# Get the active worksheet (default: first sheet)
ws = wb.active

# ==================================================
# 3️⃣ Unmerge previously merged cells
# ==================================================
# Example: Unmerge cells that were merged from B2 to D2
# IMPORTANT:
# - The range must exactly match the original merged range
# - Otherwise, openpyxl may raise an error
ws.unmerge_cells("B2:D2")

# ==================================================
# 4️⃣ Save the result as a new Excel file
# ==================================================
# Saving as a new file is recommended to avoid overwriting the original
wb.save("sample_unmerge.xlsx")

print("✅ Cells successfully unmerged and saved as sample_unmerge.xlsx")
