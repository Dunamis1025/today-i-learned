"""
📌 Merging Cells in Excel (openpyxl)

This script demonstrates how to merge multiple Excel cells into a single cell
using the openpyxl library.

What this file covers:
- Merging a range of cells into one
- Writing a value into a merged cell
- Understanding how merged cells actually work in Excel

🎯 Purpose
- Learning-focused example for Excel layout design
- Useful for creating titles, headers, and section labels
- Foundation for report-style Excel automation
"""

# ==================================================
# 1️⃣ Import required library
# ==================================================
from openpyxl import Workbook


# ==================================================
# 2️⃣ Create a new Excel workbook
# ==================================================
wb = Workbook()
ws = wb.active  # Use the default active worksheet


# ==================================================
# 3️⃣ Merge cells
# ==================================================
# Merge cells from B2 to D2
# → B2, C2, and D2 become one merged cell
ws.merge_cells("B2:D2")


# ==================================================
# 4️⃣ Assign a value to the merged cell
# ==================================================
# Only the top-left cell (B2) is a real cell
# Values must be written to the first cell of the merged range
ws["B2"].value = "Merged Cell"


# ==================================================
# 5️⃣ Save the Excel file
# ==================================================
wb.save("sample_merge.xlsx")
