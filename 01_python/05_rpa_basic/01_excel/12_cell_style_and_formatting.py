"""
📌 Excel Cell Styling & Formatting (openpyxl)

This file demonstrates how to apply various styles to Excel cells
using the openpyxl library.

What this file covers:
- Setting column widths and row heights
- Changing font color, size, and font family
- Applying text styles (bold, italic, underline, strike-through)
- Adding borders to cells

🎯 Purpose
- This is a learning note, not a production automation script
- The goal is to understand how Excel styling works in openpyxl
- Useful later for reports, summaries, and formatted output files
"""

# ==================================================
# 1️⃣ Import required libraries
# ==================================================
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side


# ==================================================
# 2️⃣ Load Excel workbook
# ==================================================
# Load an existing Excel file (must be in the same directory)
wb = load_workbook("sample.xlsx")
ws = wb.active  # Use the active worksheet


# ==================================================
# 3️⃣ Access target cells
# ==================================================
# Header cells
# A1: Number / B1: English / C1: Math
a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]


# ==================================================
# 4️⃣ Set column width and row height
# ==================================================
# Set width of column A
ws.column_dimensions["A"].width = 5

# Set height of the first row
ws.row_dimensions[1].height = 50


# ==================================================
# 5️⃣ Apply font styles
# ==================================================
# A1 cell:
# - Red text
# - Italic
# - Bold
a1.font = Font(
    color="FF0000",
    italic=True,
    bold=True
)

# B1 cell:
# - Purple text
# - Arial font
# - Strike-through text
b1.font = Font(
    color="CC33FF",
    name="Arial",
    strike=True
)

# C1 cell:
# - Blue text
# - Font size 20
# - Underline
c1.font = Font(
    color="0000FF",
    size=20,
    underline="single"
)


# ==================================================
# 6️⃣ Apply cell borders
# ==================================================
# Create a thin border style
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Apply the same border to each header cell
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border


# ==================================================
# 7️⃣ Save the styled Excel file
# ==================================================
# Save as a new file to preserve the original
wb.save("sample_style.xlsx")
