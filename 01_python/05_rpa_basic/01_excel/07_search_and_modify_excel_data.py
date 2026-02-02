"""
📌 Searching and Modifying Existing Excel Data (openpyxl)

This file demonstrates how to:
1) Load an existing Excel file
2) Search through data rows using conditions
3) Modify header values in the first row
4) Save the result as a new file without overwriting the original

❗ Purpose
- Focus on understanding how to read and modify existing Excel data
- Practice conditional logic with iter_rows()
- Keep the original file safe by saving a modified copy
"""

# ==================================================
# 1️⃣ Import required function
# ==================================================
from openpyxl import load_workbook


# ==================================================
# 2️⃣ Load an existing Excel file
# ==================================================
# Load sample.xlsx and create a Workbook object
wb = load_workbook("sample.xlsx")

# Get the active worksheet
ws = wb.active


# ==================================================
# 3️⃣ Iterate over data rows and apply conditions
# ==================================================
"""
ws.iter_rows(min_row=2)
- Starts iteration from the 2nd row
- Commonly used to skip the header row

Each 'row' is a tuple of Cell objects.
Example structure for this sheet:
row[0] -> ID
row[1] -> English score
row[2] -> Math score
"""

for row in ws.iter_rows(min_row=2):
    # Convert the English score to int for safe comparison
    if int(row[1].value) > 80:
        # row[0].value contains the student ID
        print(row[0].value, "is an English genius")


# ==================================================
# 4️⃣ Modify header values in the first row
# ==================================================
"""
ws.iter_rows(max_row=1)
- Limits iteration to the first row only
- Useful for modifying column headers
"""

for row in ws.iter_rows(max_row=1):
    for cell in row:
        # Replace the header name "English" with "Computer"
        if cell.value == "English":
            cell.value = "Computer"


# ==================================================
# 5️⃣ Save the modified workbook as a new file
# ==================================================
"""
Important:
- Do NOT overwrite the original file
- Save the modified version as a new file
"""
wb.save("sample_modified.xlsx")
