"""
📌 Working with Excel Cell Ranges and Iteration (openpyxl)

This file is a learning note that summarizes how to:
- Select a specific column or multiple columns
- Select a specific row or a range of rows
- Understand why openpyxl returns tuples
- Use iter_rows() and iter_cols() effectively

❗ Purpose
- Not focused on automation output
- Focused on understanding Excel structure and iteration logic
- Written so future-me can quickly recall the concepts
"""

# ==================================================
# 1️⃣ Import required libraries
# ==================================================
from openpyxl import Workbook
from random import randint


# ==================================================
# 2️⃣ Create Workbook and Worksheet
# ==================================================
wb = Workbook()
ws = wb.active


# ==================================================
# 3️⃣ Add data row by row
# ==================================================
# Header row
ws.append(["ID", "English", "Math"])

# Data rows (2nd row to 11th row)
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])


# ==================================================
# 4️⃣ Select a specific column
# ==================================================
# ws["B"] returns all cells in column B
# The result is a tuple of Cell objects
col_B = ws["B"]

# Example: print English scores only
# for cell in col_B:
#     print(cell.value)


# ==================================================
# 5️⃣ Select a range of columns
# ==================================================
# ws["B:C"] returns columns B to C
# The result is a tuple of tuples (columns)
col_range = ws["B:C"]

# Structure example:
# (
#   (B1, B2, B3, ...),
#   (C1, C2, C3, ...)
# )

# for cols in col_range:
#     for cell in cols:
#         print(cell.value)


# ==================================================
# 6️⃣ Select a specific row
# ==================================================
# ws[1] returns the first row
row_title = ws[1]

# for cell in row_title:
#     print(cell.value)


# ==================================================
# 7️⃣ Select a range of rows
# ==================================================
# ws[2:6] returns rows from 2 to 6
row_range = ws[2:6]

# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()


# ==================================================
# 8️⃣ Understanding tuples in openpyxl
# ==================================================
"""
Most selections in openpyxl return tuples.

Why tuples?
- Ordered
- Immutable (safe, read-only)
- Ideal for iteration

Key idea:
👉 openpyxl gives cell collections as tuples
👉 We usually iterate over them, not modify them
"""

# All rows
# print(tuple(ws.rows))

# All columns
# print(tuple(ws.columns))

# Print column headers only
# for column in ws.columns:
#     print(column[0].value)


# ==================================================
# 9️⃣ Using iter_rows()
# ==================================================
"""
iter_rows() is used to iterate row by row with conditions.

Typical use cases:
- Skip header rows
- Process specific row/column ranges
"""

# Rows 2 to 11, Columns 2 to 3 (English, Math)
for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
    print(row[0].value, row[1].value)


# ==================================================
# 🔟 Using iter_cols()
# ==================================================
"""
iter_cols() is used to iterate column by column.

Typical use cases:
- Column-based calculations (averages, totals)
- Data analysis by subject
"""

# Rows 1 to 5, Columns 1 to 3
for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print([cell.value for cell in col])


# ==================================================
# 11️⃣ Save the Excel file
# ==================================================
wb.save("sample.xlsx")
