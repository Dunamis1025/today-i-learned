"""
📌 Excel Formulas and DateTime Handling (openpyxl)

This script demonstrates how to write Excel formulas and date values
into cells using openpyxl.

What this file covers:
- Inserting today's date using Python's datetime module
- Writing Excel formulas (SUM, AVERAGE) into cells
- Using cell ranges inside formulas
- Understanding that Excel, not Python, performs the actual calculation

🎯 Purpose
- Learning-focused example for Excel automation
- To understand how formula-based calculations work in Excel files
- Foundation for automated reports and calculation sheets
"""

# ==================================================
# 1️⃣ Import required libraries
# ==================================================
import datetime
from openpyxl import Workbook


# ==================================================
# 2️⃣ Create a new Excel workbook
# ==================================================
wb = Workbook()
ws = wb.active  # Use the default active worksheet


# ==================================================
# 3️⃣ Insert date and time
# ==================================================
# Insert today's date and time into cell A1
# datetime.datetime.today() returns the current date and time
ws["A1"] = datetime.datetime.today()


# ==================================================
# 4️⃣ Write basic Excel formulas
# ==================================================
# Excel formulas must be written as strings starting with "="

# 1 + 2 + 3 = 6
ws["A2"] = "=SUM(1, 2, 3)"

# (1 + 2 + 3) / 3 = 2
ws["A3"] = "=AVERAGE(1, 2, 3)"


# ==================================================
# 5️⃣ Use cell ranges in formulas
# ==================================================
# Insert numeric values
ws["A4"] = 10
ws["A5"] = 20

# Sum values from A4 to A5 → 30
ws["A6"] = "=SUM(A4:A5)"


# ==================================================
# 6️⃣ Save the Excel file
# ==================================================
# Save the workbook containing formulas
wb.save("sample_formula.xlsx")
