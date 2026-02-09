"""
📌 Cell Alignment, Conditional Formatting, and Freeze Panes (openpyxl)

This script demonstrates how to improve Excel readability using openpyxl by:
- Aligning cell contents horizontally and vertically
- Applying conditional styling based on cell values
- Highlighting high scores with background and font colors
- Freezing rows and columns to keep headers visible while scrolling

🎯 Purpose
- Learning-focused example (not production automation)
- To understand how to make Excel reports easier to read
- Useful for score sheets, reports, and summary tables
"""

# ==================================================
# 1️⃣ Import required libraries
# ==================================================
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment


# ==================================================
# 2️⃣ Load Excel workbook
# ==================================================
# Load an existing Excel file from the same directory
wb = load_workbook("sample.xlsx")
ws = wb.active


# ==================================================
# 3️⃣ Access header cells
# ==================================================
# A1: Number / B1: English / C1: Math
a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]


# ==================================================
# 4️⃣ Set column width and row height
# ==================================================
# Narrow the number column
ws.column_dimensions["A"].width = 5

# Increase header row height
ws.row_dimensions[1].height = 50


# ==================================================
# 5️⃣ Apply header font styles
# ==================================================
a1.font = Font(color="FF0000", italic=True, bold=True)
b1.font = Font(color="CC33FF", name="Arial", strike=True)
c1.font = Font(color="0000FF", size=20, underline="single")


# ==================================================
# 6️⃣ Apply borders to header cells
# ==================================================
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

a1.border = thin_border
b1.border = thin_border
c1.border = thin_border


# ==================================================
# 7️⃣ Alignment and conditional formatting
# ==================================================
# Iterate through all rows and cells
for row in ws.rows:
    for cell in row:
        # Center-align all cells (horizontal & vertical)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # Skip the first column (student number)
        if cell.column == 1:
            continue

        # Apply style only if the value is an integer score above 90
        if isinstance(cell.value, int) and cell.value > 90:
            # Green background for high scores
            cell.fill = PatternFill(
                fgColor="00FF00",
                fill_type="solid"
            )

            # Red font color to emphasize the score
            cell.font = Font(color="FF0000")


# ==================================================
# 8️⃣ Freeze panes
# ==================================================
# Freeze panes at B2
# - Keeps the header row visible when scrolling down
# - Keeps the number column visible when scrolling right
ws.freeze_panes = "B2"


# ==================================================
# 9️⃣ Save the file
# ==================================================
# Save as a new Excel file
wb.save("sample_style.xlsx")
