"""
📌 Reading Excel Formula Results with data_only (openpyxl)

This script demonstrates the difference between:
- Reading Excel formulas themselves
- Reading the calculated results of those formulas

It focuses on understanding the `data_only` option in openpyxl
and why formula cells may return None.
"""

# ==================================================
# 1️⃣ Import required library
# ==================================================
from openpyxl import load_workbook


# ==================================================
# 2️⃣ Load workbook WITHOUT data_only
# ==================================================
# Default behavior: data_only=False
# → Reads the formula text itself (e.g. "=SUM(A1:A5)")
#
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active
#
# print("📌 data_only=False (reading formulas)")
# for row in ws.values:
#     for cell in row:
#         print(cell)

"""
Expected output:
- '=SUM(1, 2, 3)'
- '=AVERAGE(1, 2, 3)'
- '=SUM(A4:A5)'

This means openpyxl is reading the formula strings,
not the calculated values.
"""


# ==================================================
# 3️⃣ Load workbook WITH data_only=True
# ==================================================
# data_only=True
# → Reads the calculated result stored by Excel
wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active


# ==================================================
# 4️⃣ Read cell values
# ==================================================
print("📌 data_only=True (reading calculated values)")

for row in ws.values:
    for cell in row:
        print(cell)

"""
⚠️ Very important concept:

1️⃣ openpyxl does NOT evaluate formulas
- Formula calculation is done by Excel itself

2️⃣ Excel must have calculated the file at least once
- The file needs to be opened and saved in Excel

3️⃣ If Excel has never calculated the formula,
- openpyxl has no stored result to read
- The value will be None

➡️ None does NOT mean an error
➡️ It means "Excel has not calculated this yet"
"""


# ==================================================
# 5️⃣ Final summary
# ==================================================
"""
✔ data_only=False (default)
- Reads formula text
- Example: '=SUM(A4:A5)'

✔ data_only=True
- Reads calculated result
- Example: 30
- Returns None if Excel has not calculated the formula

📌 Typical automation workflow:
1) Generate Excel file with formulas using Python
2) Open the file in Excel to trigger calculation
3) Read results using data_only=True
"""
