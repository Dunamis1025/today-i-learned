"""
📌 openpyxl Basics - Working with Excel Cells using Python (Learning Notes)

In this file, we practice the core fundamentals of handling Excel cells with openpyxl.

What you will learn:
1️⃣ Creating a new Excel workbook and worksheet
2️⃣ Renaming the active worksheet
3️⃣ Writing values to specific cells (A1-style access)
4️⃣ Understanding the difference between a Cell object and its value
5️⃣ Accessing cells using row and column numbers
6️⃣ Writing values using loops
7️⃣ Filling cells sequentially using an index variable

⚠️ Purpose
- This is NOT for production automation
- This is a learning note to understand Excel's cell structure
- Focus is on clarity and fundamentals, not performance
"""

# ==================================================
# 1️⃣ Create a Workbook (Excel file)
# ==================================================
from openpyxl import Workbook

# Create a new Excel workbook
wb = Workbook()

# Get the active worksheet
ws = wb.active

# Rename the worksheet
ws.title = "NadoSheet"


# ==================================================
# 2️⃣ Writing values to cells (A1-style access)
# ==================================================
# Write values into column A
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

# Write values into column B
ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6


# ==================================================
# 3️⃣ Cell object vs Cell value
# ==================================================
# Printing a cell directly shows the Cell object itself
print(ws["A1"])           # <Cell 'NadoSheet'.A1>

# To get the actual value inside the cell, use .value
print(ws["A1"].value)     # 1

# If a cell has no value, .value returns None
print(ws["A10"].value)    # None


# ==================================================
# 4️⃣ Accessing cells using row and column numbers
# ==================================================
# Excel structure reminder:
# row    -> 1, 2, 3, ...
# column -> A=1, B=2, C=3, ...

# column=1, row=1  -> A1
print(ws.cell(column=1, row=1).value)   # Same as ws["A1"].value

# column=2, row=1  -> B1
print(ws.cell(column=2, row=1).value)   # Same as ws["B1"].value


# ==================================================
# 5️⃣ Writing a value using cell() and storing the cell
# ==================================================
# Write value 10 into C1 and store the Cell object
c = ws.cell(column=3, row=1, value=10)  # Same as ws["C1"] = 10

# Print the value stored in the cell
print(c.value)  # 10


# ==================================================
# 6️⃣ Filling cells using loops
# ==================================================
from random import *

# Index variable for sequential values
index = 1

# Loop through rows (1 to 10)
for row in range(1, 11):
    # Loop through columns (1 to 10)
    for column in range(1, 11):

        # Option 1: Fill with random numbers (0~100)
        # ws.cell(row=row, column=column, value=randint(0, 100))

        # Option 2: Fill sequential numbers using index
        ws.cell(row=row, column=column, value=index)

        index += 1


# ==================================================
# 7️⃣ Save the Excel file
# ==================================================
wb.save("sample.xlsx")
