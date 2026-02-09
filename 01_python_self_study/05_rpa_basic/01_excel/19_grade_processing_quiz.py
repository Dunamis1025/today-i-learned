"""
📌 [Quiz] Student Grade Processing Automation (openpyxl)

This script demonstrates how to process student grade data using openpyxl.

✔ What this script does
1. Updates all Quiz 2 scores to full marks (10 points) due to an error in the quiz
2. Calculates the total score and writes it to column H using the SUM formula
3. Assigns letter grades (A/B/C/D) based on the total score
4. Assigns an F grade to students with attendance less than 5, regardless of total score
5. Saves the final results to an Excel file (scores.xlsx)

📂 Output file
- scores.xlsx
"""

# ==================================================
# 1️⃣ Import required library
# ==================================================
from openpyxl import Workbook


# ==================================================
# 2️⃣ Create a new Excel workbook and worksheet
# ==================================================
wb = Workbook()        # Create a new Excel workbook
ws = wb.active        # Select the active worksheet


# ==================================================
# 3️⃣ Add header row
# ==================================================
ws.append(("Student ID", "Attendance", "Quiz 1", "Quiz 2",
           "Midterm", "Final Exam", "Project"))


# ==================================================
# 4️⃣ Student grade data
# ==================================================
scores = [
    (1, 10, 8, 5, 14, 26, 12),
    (2, 7, 3, 7, 15, 24, 18),
    (3, 9, 5, 8, 8, 12, 4),
    (4, 7, 8, 7, 17, 21, 18),
    (5, 7, 8, 7, 16, 25, 15),
    (6, 3, 5, 8, 8, 17, 0),
    (7, 4, 9, 10, 16, 27, 18),
    (8, 6, 6, 6, 15, 19, 17),
    (9, 10, 10, 9, 19, 30, 19),
    (10, 9, 8, 8, 20, 25, 20)
]

# Append student data to the worksheet
for s in scores:
    ws.append(s)


# ==================================================
# 5️⃣ Update Quiz 2 scores to full marks
# ==================================================
# Column D corresponds to "Quiz 2"
for idx, cell in enumerate(ws["D"]):
    if idx == 0:          # Skip header row
        continue
    cell.value = 10      # Set Quiz 2 score to 10


# ==================================================
# 6️⃣ Add headers for total score and grade
# ==================================================
ws["H1"] = "Total Score"
ws["I1"] = "Grade"


# ==================================================
# 7️⃣ Calculate total scores and assign grades
# ==================================================
# Data rows start from row 2
for idx, score in enumerate(scores, start=2):

    # Adjust total score by replacing Quiz 2 with full marks
    total_score = sum(score[1:]) - score[3] + 10

    # Write Excel SUM formula to column H
    # Example: =SUM(B2:G2)
    ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)

    # Determine grade based on total score
    grade = None
    if total_score >= 90:
        grade = "A"
    elif total_score >= 80:
        grade = "B"
    elif total_score >= 70:
        grade = "C"
    else:
        grade = "D"

    # Students with attendance less than 5 receive an F
    if score[1] < 5:
        grade = "F"

    # Write grade to column I
    ws.cell(row=idx, column=9).value = grade


# ==================================================
# 8️⃣ Save the final Excel file
# ==================================================
wb.save("scores.xlsx")
