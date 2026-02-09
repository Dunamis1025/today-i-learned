"""
📌 RPA Basics - Excel Automation (openpyxl)
02) Sheet operations: create, rename, reorder, copy, and style

This script demonstrates how to manage worksheets (sheets)
inside an Excel workbook using openpyxl.

What you'll learn in this file:
1) Create new sheets with default or custom names
2) Control the position (index) where a sheet is created
3) Change the sheet tab color using RGB hex values
4) Access sheets by name (dictionary-style access)
5) List all sheet names in a workbook
6) Copy an existing sheet
7) Write data to a specific cell (e.g., A1)
"""

from openpyxl import Workbook

# ==================================================
# 1️⃣ Create a new workbook
# ==================================================
wb = Workbook()

# ==================================================
# 2️⃣ Create a new sheet with the default name
# ==================================================
# If no name is provided, openpyxl assigns a default name automatically.
ws = wb.create_sheet()

# Rename the sheet to something more meaningful
ws.title = "MySheet"

# ==================================================
# 3️⃣ Change the sheet tab color
# ==================================================
# You can change the sheet tab color using an RGB hex code.
# Color codes can be found here:
# https://www.w3schools.com/colors/colors_picker.asp
ws.sheet_properties.tabColor = "ff66ff"  # Pink color

# ==================================================
# 4️⃣ Create a sheet with a specific name
# ==================================================
ws1 = wb.create_sheet("YourSheet")

# ==================================================
# 5️⃣ Create a sheet at a specific position (index)
# ==================================================
# The index starts from 0.
# This creates "NewSheet" as the third sheet in the workbook.
ws2 = wb.create_sheet("NewSheet", 2)

# ==================================================
# 6️⃣ Access a sheet by name
# ==================================================
# Sheets can be accessed like dictionary keys.
new_ws = wb["NewSheet"]

# ==================================================
# 7️⃣ List all sheet names in the workbook
# ==================================================
print(wb.sheetnames)

# ==================================================
# 8️⃣ Copy a worksheet
# ==================================================
# First, write some data to the original sheet.
new_ws["A1"] = "Test"

# Copy the entire worksheet.
copied_ws = wb.copy_worksheet(new_ws)

# Rename the copied sheet.
copied_ws.title = "Copied Sheet"

# ==================================================
# 9️⃣ Save the workbook
# ==================================================
wb.save("sample.xlsx")
