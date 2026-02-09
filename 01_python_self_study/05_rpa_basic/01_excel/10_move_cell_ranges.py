"""
📌 Moving Cell Ranges in Excel (move_range) - openpyxl

This file demonstrates how to:
- Move a specific range of cells to a new position
- Shift data vertically and horizontally using rows and cols
- Understand the difference between moving and copying data

❗ Purpose
- Learn how to rearrange Excel layouts programmatically
- Practice structural manipulation using move_range()
- Keep this file as a reference for Excel automation
"""

# ==================================================
# 1️⃣ Load an existing Excel file
# ==================================================
from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active


# ==================================================
# 2️⃣ Example 1: Reordering columns
# ==================================================
"""
Original layout:
ID | English | Math

Target layout:
ID | Korean | English | Math

Steps:
- Move the range B1:C11 one column to the right
- Insert a new subject name in cell B1
"""

# ws.move_range("B1:C11", rows=0, cols=1)
# ws["B1"].value = "Korean"


# ==================================================
# 3️⃣ Example 2: Moving a range by rows and columns
# ==================================================
"""
This example moves the range C1:C11:
- Down by 5 rows
- Left by 1 column

rows > 0  → move downward
cols < 0  → move left
"""

ws.move_range("C1:C11", rows=5, cols=-1)


# ==================================================
# 4️⃣ Save the modified workbook
# ==================================================
wb.save("sample_korean.xlsx")
